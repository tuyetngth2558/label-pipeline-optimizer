#!/usr/bin/env python3
"""
Pre-Submit Validator — scan Excel trước khi nộp.

  python validator.py -f outputs/annotation_output.xlsx
  python validator.py -f batch.xlsx --strict   # bắt buộc intern đã review
"""
from __future__ import annotations

import argparse
import re
import sys
from urllib.parse import urlparse

import openpyxl

from modules.claim_constraints import PLACEHOLDER_G

VALID_STATUSES = {
    "XAC NHAN", "LECH", "MAU THUAN", "OUTDATED", "KHONG TIM THAY", "BO QUA",
}
NOTES_KEYS = ("SF=", "SC=", "HR=", "SQ=", "TXT=")
NUM_COLS = 18


def _is_full_url(url: str) -> bool:
    if not url or not url.startswith("http"):
        return False
    parsed = urlparse(url)
    return bool(parsed.netloc and parsed.path and parsed.path != "/")


def validate_row(row_num: int, row: tuple, strict: bool = False) -> list[str]:
    issues = []
    vals = list(row) + [None] * (NUM_COLS - len(row))
    vals = vals[:NUM_COLS]
    claim_num = vals[0] or row_num - 4

    if not vals[5]:
        return issues

    status = str(vals[6] or "").strip().upper()
    url = str(vals[7] or "").strip()
    sc = vals[9]
    hr = vals[10]
    notes = str(vals[12] or "")
    evidence = str(vals[15] or "").strip() if len(vals) > 15 else ""
    url_load = str(vals[16] or "").strip().upper() if len(vals) > 16 else ""
    intern_rev = str(vals[17] or "").strip().upper() if len(vals) > 17 else ""

    if strict and PLACEHOLDER_G in str(vals[6] or ""):
        issues.append(f"Claim #{claim_num} — Cột G: vẫn placeholder — intern chưa fact-check")

    if strict and (not intern_rev or intern_rev == "N"):
        issues.append(f"Claim #{claim_num} — Cột R: intern_reviewed chưa = Y")

    if PLACEHOLDER_G in str(vals[6] or "") and not strict:
        return issues  # pre-label file — chỉ warn khi --strict

    if status and status not in VALID_STATUSES:
        issues.append(f"Claim #{claim_num} — Cột G: status không hợp lệ '{status}'")

    if status != "BO QUA" and status and not url:
        issues.append(f"Claim #{claim_num} — Cột H: thiếu URL (status={status})")

    if url and not _is_full_url(url):
        issues.append(f"Claim #{claim_num} — Cột H: URL không đầy đủ (thiếu path)")

    for label, val, col in (("SC", sc, "J"), ("HR", hr, "K")):
        try:
            v = float(val) if val is not None else 0.0
            if v == 0.0:
                issues.append(f"Claim #{claim_num} — Cột {col}: {label}=0.00 chưa điền")
        except (TypeError, ValueError):
            issues.append(f"Claim #{claim_num} — Cột {col}: {label} không phải số")

    if notes:
        line_count = len([ln for ln in notes.splitlines() if ln.strip()])
        if line_count < 5:
            issues.append(f"Claim #{claim_num} — Cột M: Notes chỉ {line_count} dòng (cần ≥5)")
        notes_u = notes.upper().replace(":", "=")
        for key in NOTES_KEYS:
            if key not in notes_u and f"DRAFT_{key}" not in notes_u:
                issues.append(f"Claim #{claim_num} — Cột M: thiếu {key.rstrip('=')}=")
    else:
        issues.append(f"Claim #{claim_num} — Cột M: thiếu Notes")

    try:
        hr_v = float(hr) if hr is not None else None
        sc_v = float(sc) if sc is not None else None
        if hr_v is not None and status == "XAC NHAN" and hr_v < 0.35:
            issues.append(
                f"Claim #{claim_num} — HR={hr_v:.2f} thấp nhưng Status=XAC NHAN → kiểm tra lại"
            )
        if sc_v is not None and status == "XAC NHAN" and sc_v > 0.7 and strict:
            if url_load and url_load not in ("Y", "YES", "TRUE", "1"):
                issues.append(
                    f"Claim #{claim_num} — SC cao nhưng URL Load OK={url_load} → kiểm tra nguồn"
                )
        if status == "LECH" and not url:
            issues.append(f"Claim #{claim_num} — Status=LECH nhưng cột H trống")
    except (TypeError, ValueError):
        pass

    if strict and status == "XAC NHAN":
        if not evidence:
            issues.append(f"Claim #{claim_num} — Cột P: thiếu Evidence Quote (bắt buộc với XAC NHAN)")
        if url_load not in ("Y", "YES", "TRUE", "1"):
            issues.append(f"Claim #{claim_num} — Cột Q: url_load_ok phải Y khi XAC NHAN")

    return issues


def validate_workbook(
    path: str,
    sheet: str = "Annotation",
    from_row: int = 5,
    strict: bool = False,
) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return {"ok": False, "errors": [f"Không có sheet '{sheet}'"], "claims": 0}

    ws = wb[sheet]
    all_issues: list[str] = []
    claim_count = 0
    placeholder_count = 0

    for r in range(from_row, ws.max_row + 1):
        row = tuple(ws.cell(row=r, column=c).value for c in range(1, NUM_COLS + 1))
        if not row[5]:
            continue
        claim_count += 1
        if PLACEHOLDER_G in str(row[6] or ""):
            placeholder_count += 1
        all_issues.extend(validate_row(r, row, strict=strict))

    wb.close()

    if strict and placeholder_count > 0:
        all_issues.insert(
            0,
            f"❌ Còn {placeholder_count} claim chưa fact-check (cột G placeholder)",
        )

    return {
        "ok": len(all_issues) == 0,
        "errors": all_issues,
        "claims": claim_count,
        "placeholder_count": placeholder_count,
    }


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser(description="Vivipedia Pre-Submit Validator")
    ap.add_argument("--file", "-f", required=True)
    ap.add_argument("--sheet", default="Annotation")
    ap.add_argument("--from-row", type=int, default=5)
    ap.add_argument(
        "--strict",
        action="store_true",
        help="Chế độ nộp bài: không placeholder, intern_reviewed=Y, evidence bắt buộc",
    )
    args = ap.parse_args()

    mode_label = "STRICT (sẵn sàng nộp)" if args.strict else "STANDARD"
    print(f"\n📋 KIỂM TRA [{mode_label}]: {args.file}\n")
    result = validate_workbook(args.file, args.sheet, args.from_row, strict=args.strict)

    if not result["claims"]:
        print("⚠ Không có claim nào từ dòng", args.from_row)
        sys.exit(1)

    if not args.strict and result.get("placeholder_count", 0) == result["claims"]:
        print("ℹ️  File pre-label: tất cả claim còn placeholder — dùng --strict sau khi intern review.\n")

    for err in result["errors"]:
        prefix = "❌" if err.startswith("❌") or "placeholder" in err.lower() or "thiếu" in err.lower() else "⚠️ "
        print(f"{prefix} {err}")

    n_err = len(result["errors"])
    print(f"\n📊 {result['claims']} claim | {n_err} vấn đề")
    if n_err:
        print("❌ Chưa đạt — sửa trước khi nộp\n")
        sys.exit(1)
    print("\n✅ File đạt chuẩn.\n")


if __name__ == "__main__":
    main()
