"""
excel_writer.py — Excel annotation v10 + cột REVIEW (P–R).
"""
import os
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "outputs", "annotation_output.xlsx")

NUM_COLS = 18  # A–R

_NAVY = PatternFill("solid", fgColor="FF1F3864")
_RED = PatternFill("solid", fgColor="FFC00000")
_BLUE = PatternFill("solid", fgColor="FF2E75B6")
_GRAY = PatternFill("solid", fgColor="FF374151")
_GREEN = PatternFill("solid", fgColor="FF1D6F42")
_DATABG = PatternFill("solid", fgColor="FFEBF8FF")
_STTBG = PatternFill("solid", fgColor="FFF2F2F2")

_COL_WIDTHS = [5, 20, 16, 20, 12, 48, 18, 32, 9, 9, 9, 9, 40, 11, 14, 36, 10, 14]

_COL_LABELS = [
    "#", "Article / Page Title", "Domain", "Sub-domain", "Sub-domain\nID",
    "Claim (block nguyên văn)", "Fact-check\nStatus", "Fact-check\nSource URL",
    "Source\nFidelity\n(SF)", "Source\nCoverage\n(SC)",
    "Hallucination\nRate (HR)\n(inv.)", "Source\nQuality\n(SQ)",
    "Annotator Notes", "Annotator\nID", "Date",
    "Evidence\nQuote", "URL\nLoad OK", "Intern\nReviewed",
]

_TEMPLATE_VALS = [
    "auto", "[article title]", "[select domain]", "[select sub-domain]", "[sub_id]",
    "[paste claim block]",
    "[XAC NHAN / ... / [INTERN: chưa fact-check]]",
    "[https://...]", "0-1", "0-1", "0-1", "0-1",
    "[SF=... / SC=... / HR=... / SQ=... / TXT=...]",
    "[ANT-xx]", "[date]",
    "[trích dẫn từ URL ≤200 ký tự]", "Y/N", "N → Y sau review",
]


def _col_fill(col: int) -> PatternFill:
    if col in {7, 8}:
        return _RED
    if col in {9, 10, 11, 12}:
        return _BLUE
    if col in {13, 14, 15}:
        return _GRAY
    if col in {16, 17, 18}:
        return _GREEN
    return _NAVY


def _hdr(cell, value: str, fill: PatternFill, size: int = 8):
    cell.value = value
    cell.font = Font(bold=True, color="FFFFFFFF", name="Arial", size=size)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _data(cell, value, bold=False, size=9, color="FF000000",
          fill: PatternFill | None = None, h_align="left"):
    cell.value = value
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=h_align, vertical="top", wrap_text=True)
    if fill:
        cell.fill = fill


def _build_header(ws):
    last_col = get_column_letter(NUM_COLS)
    ws.merge_cells(f"A1:{last_col}1")
    _hdr(
        ws["A1"],
        "RAG ANNOTATION v10 — SF/SC/HR/SQ | Evidence + Intern Review (P–R)",
        _NAVY,
        size=11,
    )
    ws.row_dimensions[1].height = 18

    ws.merge_cells("A2:F2")
    ws.merge_cells("G2:H2")
    ws.merge_cells("I2:L2")
    ws.merge_cells("M2:O2")
    ws.merge_cells("P2:R2")
    _hdr(ws["A2"], "IDENTITY", _NAVY, size=9)
    _hdr(ws["G2"], "FACT-CHECK", _RED, size=9)
    _hdr(ws["I2"], "METRICS", _BLUE, size=9)
    _hdr(ws["M2"], "ANNOTATION INFO", _GRAY, size=9)
    _hdr(ws["P2"], "REVIEW / EVIDENCE", _GREEN, size=9)
    ws.row_dimensions[2].height = 18

    for col, label in enumerate(_COL_LABELS, 1):
        _hdr(ws.cell(row=3, column=col), label, _col_fill(col), size=8)
    ws.row_dimensions[3].height = 36

    for col, val in enumerate(_TEMPLATE_VALS, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = Font(name="Arial", size=7, italic=True, color="FF888888")
        cell.fill = _DATABG if col > 1 else _STTBG
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C5"


def _ensure_header(ws):
    """Nâng cấp file Excel cũ (15 cột) lên 18 cột."""
    if ws.max_column >= NUM_COLS:
        lbl = ws.cell(row=3, column=16).value
        if lbl and "Evidence" in str(lbl):
            return
    _build_header(ws)


def get_or_create_workbook() -> tuple[openpyxl.Workbook, str]:
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    if os.path.exists(OUTPUT_PATH):
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        if "Annotation" not in wb.sheetnames:
            ws = wb.create_sheet("Annotation")
            _build_header(ws)
        else:
            _ensure_header(wb["Annotation"])
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "Annotation"
        _build_header(wb.active)
    return wb, OUTPUT_PATH


def _next_stt(ws) -> int:
    for r in range(ws.max_row, 4, -1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            try:
                return int(val) + 1
            except (TypeError, ValueError):
                pass
    return 1


def append_rows(rows: list[list], log_fn=print) -> str:
    """
    rows: mỗi list 18 phần tử [A..R]. STT (A) có thể "".
    """
    wb, path = get_or_create_workbook()
    ws = wb["Annotation"]

    start_row = max(ws.max_row + 1, 5)
    if ws.max_row >= 5:
        last = ws.max_row
        if ws.cell(row=last, column=1).value is None and ws.cell(row=last, column=6).value in (None, ""):
            start_row = last

    stt = _next_stt(ws)

    for i, row_data in enumerate(rows):
        r = start_row + i
        data_row = list(row_data) + [""] * (NUM_COLS - len(row_data))
        data_row = data_row[:NUM_COLS]
        data_row[0] = stt + i

        for col, val in enumerate(data_row, 1):
            cell = ws.cell(row=r, column=col)
            bg = _STTBG if col == 1 else _DATABG
            if col in {9, 10, 11, 12}:
                _data(cell, val, bold=True, size=10, fill=bg, h_align="center")
            elif col == 8:
                _data(cell, val, size=8, color="FF0563C1", fill=bg)
            elif col == 1:
                _data(cell, val, size=9, color="FF888888", fill=bg, h_align="center")
            elif col in {16, 17, 18}:
                _data(cell, val, size=8, fill=bg)
            else:
                _data(cell, val, fill=bg)
        ws.row_dimensions[r].height = 200

    try:
        wb.save(path)
    except PermissionError:
        import time
        alt = path.replace(".xlsx", f"_{int(time.time())}.xlsx")
        wb.save(alt)
        path = alt
        log_fn(f"  ⚠ File gốc đang mở → đã lưu vào: {alt}")

    log_fn(f"  Đã ghi {len(rows)} hàng (18 cột A–R) → {path}")
    return os.path.abspath(path)


def _build_article_eval_header(ws):
    ws.merge_cells("A1:N1")
    _hdr(
        ws["A1"],
        "ARTICLE EVALUATION — Relevance & Completeness (1 dòng / bài)",
        _NAVY,
        size=11,
    )
    ws.row_dimensions[1].height = 18

    labels = [
        "#", "Tên bài viết", "URL bài", "Domain", "Sub-domain",
        "Rel (0-1)", "Rel Band", "Nhận xét Relevance",
        "Comp (0-1)", "Comp Band", "Nhận xét Completeness",
        "Note", "Annotator ID", "Ngày",
    ]
    for col, label in enumerate(labels, 1):
        fill = _BLUE if col in {6, 7, 8} else (_BLUE if col in {9, 10, 11} else _NAVY)
        if col in {6, 7, 8, 9, 10, 11}:
            fill = _BLUE
        elif col in {12, 13, 14}:
            fill = _GRAY
        _hdr(ws.cell(row=3, column=col), label, fill, size=8)
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "A4"


def _next_article_stt(ws) -> int:
    for r in range(ws.max_row, 3, -1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            try:
                return int(val) + 1
            except (TypeError, ValueError):
                pass
    return 1


def append_article_evaluation(
    article: dict,
    annotator: str = "AUTO",
    today: str | None = None,
    article_url: str = "",
    log_fn=print,
) -> None:
    """Ghi 1 dòng sheet Article Evaluation (Rel/Comp cấp bài)."""
    from modules.scoring_utils import score_to_band

    wb, path = get_or_create_workbook()
    sheet_name = "Article Evaluation"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        _build_article_eval_header(ws)
    else:
        ws = wb[sheet_name]
        if ws.max_row < 3:
            _build_article_eval_header(ws)

    today = today or date.today().strftime("%Y-%m-%d")
    rel = article.get("rel", "")
    comp = article.get("comp", "")
    rel_band = article.get("rel_band") or score_to_band(rel)
    comp_band = article.get("comp_band") or score_to_band(comp)

    row = ws.max_row + 1
    if row < 4:
        row = 4
    stt = _next_article_stt(ws)

    values = [
        stt,
        article.get("title", ""),
        article_url,
        article.get("domain", ""),
        article.get("sub_domain", ""),
        rel,
        rel_band,
        article.get("rel_reason", ""),
        comp,
        comp_band,
        article.get("comp_reason", ""),
        "",
        annotator,
        today,
    ]
    for col, val in enumerate(values, 1):
        _data(ws.cell(row=row, column=col), val, fill=_DATABG)
    ws.row_dimensions[row].height = 80

    try:
        wb.save(path)
    except PermissionError:
        pass
    log_fn(f"  Article Evaluation: Rel={rel} ({rel_band}) | Comp={comp} ({comp_band})")


def write_output(stt: str, data: dict, log_fn=print) -> str:
    art = data.get("article", {})
    claims = data.get("claims", [])
    today = date.today().strftime("%Y-%m-%d")
    rows = []
    for c in claims:
        rows.append([
            "", art.get("title", ""), art.get("domain", ""),
            art.get("sub_domain", ""), art.get("sub_domain_id", ""),
            c.get("claim", ""), c.get("fact_check_status", ""),
            c.get("fact_check_source_url", ""),
            c.get("source_fidelity", ""), c.get("source_coverage", ""),
            c.get("hallucination_rate", ""), c.get("source_quality", ""),
            c.get("notes", ""), "AUTO", today,
            c.get("evidence_quote", ""), c.get("url_load_ok", "N"),
            c.get("intern_reviewed", "N"),
        ])
    return append_rows(rows, log_fn=log_fn)
